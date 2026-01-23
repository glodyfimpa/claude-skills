#!/usr/bin/env python3
"""
Generate Excel template for short-term rental business plan.
Creates a workbook with input sheet, calculations, and formatted output.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_business_plan_template(filename: str = "business_plan_template.xlsx"):
    """Create Excel template with formatted sheets and formulas"""
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets
    inputs_sheet = wb.create_sheet("Inputs")
    calc_sheet = wb.create_sheet("Calculations")
    summary_sheet = wb.create_sheet("Summary")
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    currency_format = '€#,##0.00'
    percent_format = '0.0%'
    
    # === INPUTS SHEET ===
    inputs_sheet.column_dimensions['A'].width = 30
    inputs_sheet.column_dimensions['B'].width = 15
    
    # Property Information
    inputs_sheet['A1'] = "SHORT-TERM RENTAL BUSINESS PLAN"
    inputs_sheet['A1'].font = Font(bold=True, size=14)
    
    inputs_sheet['A3'] = "Property Information"
    inputs_sheet['A3'].fill = section_fill
    inputs_sheet['A3'].font = section_font
    
    inputs_sheet['A4'] = "City"
    inputs_sheet['B4'] = "Milan"
    
    inputs_sheet['A5'] = "Zone"
    inputs_sheet['B5'] = "Navigli"
    
    inputs_sheet['A6'] = "Number of bedrooms"
    inputs_sheet['B6'] = 2
    
    # Market Data
    inputs_sheet['A8'] = "Market Data"
    inputs_sheet['A8'].fill = section_fill
    inputs_sheet['A8'].font = section_font
    
    inputs_sheet['A9'] = "Avg price per night"
    inputs_sheet['B9'] = 85
    inputs_sheet['B9'].number_format = currency_format
    
    inputs_sheet['A10'] = "Occupancy rate"
    inputs_sheet['B10'] = 0.70
    inputs_sheet['B10'].number_format = percent_format
    
    # Monthly Costs
    inputs_sheet['A12'] = "Monthly Costs"
    inputs_sheet['A12'].fill = section_fill
    inputs_sheet['A12'].font = section_font
    
    inputs_sheet['A13'] = "Monthly rent"
    inputs_sheet['B13'] = 1200
    inputs_sheet['B13'].number_format = currency_format
    
    inputs_sheet['A14'] = "Condo fees"
    inputs_sheet['B14'] = 150
    inputs_sheet['B14'].number_format = currency_format
    
    inputs_sheet['A15'] = "Utilities (gas, electricity, water)"
    inputs_sheet['B15'] = 80
    inputs_sheet['B15'].number_format = currency_format
    
    inputs_sheet['A16'] = "WiFi/Internet"
    inputs_sheet['B16'] = 30
    inputs_sheet['B16'].number_format = currency_format
    
    inputs_sheet['A17'] = "Cleaning per stay"
    inputs_sheet['B17'] = 60
    inputs_sheet['B17'].number_format = currency_format
    
    inputs_sheet['A18'] = "Supplies (toiletries, linens)"
    inputs_sheet['B18'] = 50
    inputs_sheet['B18'].number_format = currency_format
    
    inputs_sheet['A19'] = "Insurance"
    inputs_sheet['B19'] = 40
    inputs_sheet['B19'].number_format = currency_format
    
    inputs_sheet['A20'] = "Property management (%)"
    inputs_sheet['B20'] = 0.10
    inputs_sheet['B20'].number_format = percent_format
    
    # Constants
    inputs_sheet['A22'] = "Constants"
    inputs_sheet['A22'].fill = section_fill
    inputs_sheet['A22'].font = section_font
    
    inputs_sheet['A23'] = "Platform commission (Airbnb)"
    inputs_sheet['B23'] = 0.15
    inputs_sheet['B23'].number_format = percent_format
    
    inputs_sheet['A24'] = "Tax rate (cedolare secca)"
    inputs_sheet['B24'] = 0.21
    inputs_sheet['B24'].number_format = percent_format
    
    inputs_sheet['A25'] = "Avg stay length (nights)"
    inputs_sheet['B25'] = 3
    
    # === CALCULATIONS SHEET ===
    calc_sheet.column_dimensions['A'].width = 30
    calc_sheet.column_dimensions['B'].width = 15
    
    calc_sheet['A1'] = "Calculations"
    calc_sheet['A1'].font = Font(bold=True, size=12)
    
    calc_sheet['A3'] = "Nights per month"
    calc_sheet['B3'] = 30
    
    calc_sheet['A4'] = "Occupied nights"
    calc_sheet['B4'] = "=B3*Inputs!B10"
    calc_sheet['B4'].number_format = '0.0'
    
    calc_sheet['A5'] = "Number of stays"
    calc_sheet['B5'] = "=B4/Inputs!B25"
    calc_sheet['B5'].number_format = '0.0'
    
    calc_sheet['A7'] = "Gross revenue"
    calc_sheet['B7'] = "=B4*Inputs!B9"
    calc_sheet['B7'].number_format = currency_format
    
    calc_sheet['A8'] = "Net revenue (after platform fee)"
    calc_sheet['B8'] = "=B7*(1-Inputs!B23)"
    calc_sheet['B8'].number_format = currency_format
    
    calc_sheet['A10'] = "Fixed costs"
    calc_sheet['B10'] = "=SUM(Inputs!B13:B16,Inputs!B18:B19)"
    calc_sheet['B10'].number_format = currency_format
    
    calc_sheet['A11'] = "Variable costs (cleaning)"
    calc_sheet['B11'] = "=B5*Inputs!B17"
    calc_sheet['B11'].number_format = currency_format
    
    calc_sheet['A12'] = "Property management fee"
    calc_sheet['B12'] = "=B8*Inputs!B20"
    calc_sheet['B12'].number_format = currency_format
    
    calc_sheet['A13'] = "Total costs"
    calc_sheet['B13'] = "=B10+B11+B12"
    calc_sheet['B13'].number_format = currency_format
    
    calc_sheet['A15'] = "Profit before tax"
    calc_sheet['B15'] = "=B8-B13"
    calc_sheet['B15'].number_format = currency_format
    
    calc_sheet['A16'] = "Net profit (after tax)"
    calc_sheet['B16'] = "=B15*(1-Inputs!B24)"
    calc_sheet['B16'].number_format = currency_format
    
    calc_sheet['A18'] = "Profit margin"
    calc_sheet['B18'] = "=IF(B8>0,B16/B8,0)"
    calc_sheet['B18'].number_format = percent_format
    
    calc_sheet['A19'] = "Annual profit"
    calc_sheet['B19'] = "=B16*12"
    calc_sheet['B19'].number_format = currency_format
    
    calc_sheet['A20'] = "Annual ROI"
    calc_sheet['B20'] = "=IF(Inputs!B13*12>0,B19/(Inputs!B13*12),0)"
    calc_sheet['B20'].number_format = percent_format
    
    calc_sheet['A22'] = "Break-even nights"
    calc_sheet['B22'] = "=IF((Inputs!B9*(1-Inputs!B23)-(Inputs!B17/Inputs!B25))>0,B10/(Inputs!B9*(1-Inputs!B23)-(Inputs!B17/Inputs!B25)),30)"
    calc_sheet['B22'].number_format = '0'
    
    # === SUMMARY SHEET ===
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 20
    
    summary_sheet['A1'] = "BUSINESS PLAN SUMMARY"
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:B1')
    
    summary_sheet['A3'] = "Property"
    summary_sheet['A3'].fill = header_fill
    summary_sheet['A3'].font = header_font
    summary_sheet['B3'] = "=Inputs!B5&\", \"&Inputs!B4"
    summary_sheet['B3'].fill = header_fill
    summary_sheet['B3'].font = header_font
    
    summary_sheet['A5'] = "KEY METRICS"
    summary_sheet['A5'].font = section_font
    
    summary_sheet['A6'] = "Monthly Net Profit"
    summary_sheet['B6'] = "=Calculations!B16"
    summary_sheet['B6'].number_format = currency_format
    summary_sheet['B6'].font = Font(bold=True, size=12)
    
    summary_sheet['A7'] = "Annual ROI"
    summary_sheet['B7'] = "=Calculations!B20"
    summary_sheet['B7'].number_format = percent_format
    summary_sheet['B7'].font = Font(bold=True, size=12)
    
    summary_sheet['A8'] = "Profit Margin"
    summary_sheet['B8'] = "=Calculations!B18"
    summary_sheet['B8'].number_format = percent_format
    
    summary_sheet['A9'] = "Break-even (nights/month)"
    summary_sheet['B9'] = "=Calculations!B22&\"/30\""
    
    summary_sheet['A11'] = "MONTHLY BREAKDOWN"
    summary_sheet['A11'].font = section_font
    
    summary_sheet['A12'] = "Gross Revenue"
    summary_sheet['B12'] = "=Calculations!B7"
    summary_sheet['B12'].number_format = currency_format
    
    summary_sheet['A13'] = "Net Revenue (after platform)"
    summary_sheet['B13'] = "=Calculations!B8"
    summary_sheet['B13'].number_format = currency_format
    
    summary_sheet['A14'] = "Total Costs"
    summary_sheet['B14'] = "=Calculations!B13"
    summary_sheet['B14'].number_format = currency_format
    
    summary_sheet['A15'] = "  - Fixed costs"
    summary_sheet['B15'] = "=Calculations!B10"
    summary_sheet['B15'].number_format = currency_format
    
    summary_sheet['A16'] = "  - Variable costs"
    summary_sheet['B16'] = "=Calculations!B11"
    summary_sheet['B16'].number_format = currency_format
    
    summary_sheet['A17'] = "  - Property mgmt"
    summary_sheet['B17'] = "=Calculations!B12"
    summary_sheet['B17'].number_format = currency_format
    
    summary_sheet['A19'] = "DECISION CRITERIA"
    summary_sheet['A19'].font = section_font
    
    summary_sheet['A20'] = "ROI Target (min 15%)"
    summary_sheet['B20'] = "=IF(Calculations!B20>=0.15,\"✓ PASS\",\"✗ FAIL\")"
    summary_sheet['B20'].font = Font(bold=True)
    
    summary_sheet['A21'] = "Margin Target (min 20%)"
    summary_sheet['B21'] = "=IF(Calculations!B18>=0.20,\"✓ PASS\",\"✗ FAIL\")"
    summary_sheet['B21'].font = Font(bold=True)
    
    summary_sheet['A22'] = "Break-even Target (max 22 nights)"
    summary_sheet['B22'] = "=IF(Calculations!B22<=22,\"✓ PASS\",\"✗ FAIL\")"
    summary_sheet['B22'].font = Font(bold=True)
    
    # Save workbook
    wb.save(filename)
    print(f"Template created: {filename}")
    return filename


if __name__ == "__main__":
    create_business_plan_template("/home/claude/short-term-rental-analyzer/assets/business_plan_template.xlsx")
